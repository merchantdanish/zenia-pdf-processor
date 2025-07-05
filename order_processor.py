import os
import re
import csv
import datetime
import subprocess
import platform
import time
from collections import defaultdict, deque
import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Define hazmat-related keywords
HAZMAT_KEYWORDS = ["deo", "deodorant", "perfume", "parfum", "freshener", "edp", "edt", "extrait"]

# Performance optimization: Compile regex patterns once
SKU_QTY_PATTERN = re.compile(r'(T\d+)\s+(\d+)')
ORDER_ID_PATTERN = re.compile(r"Order ID:[\s]*(\d+)")
TRACKING_PATTERNS = [
    re.compile(r"Tracking\s*Number:?\s*([A-Za-z0-9]+)", re.IGNORECASE),
    re.compile(r"Tracking:?\s*([A-Za-z0-9]+)", re.IGNORECASE),
    re.compile(r"TRK\s*#:?\s*([A-Za-z0-9]+)", re.IGNORECASE)
]
QTY_TOTAL_PATTERN = re.compile(r"Qty\s+Total:\s+(\d+)", re.IGNORECASE)

def detect_page_type(page_text):
    """
    Detect if a page is a shipping label or packing slip.
    Returns: 'label' or 'packing_slip'
    """
    # Common indicators of a packing slip
    packing_slip_indicators = [
        "Order ID:",
        "Product Name",
        "SKU",
        "Qty Total:",
        "Packing Slip",
        "Ship To:"
    ]
    
    # Common indicators of a shipping label
    label_indicators = [
        "SHIP FROM:",
        "DELIVER TO:",
        "Tracking Number:",
        "Carrier:",
        "Service:"
    ]
    
    # Count indicators
    packing_slip_count = sum(1 for indicator in packing_slip_indicators if indicator in page_text)
    label_count = sum(1 for indicator in label_indicators if indicator in page_text)
    
    # If it has product listings, it's definitely a packing slip
    if "Product Name" in page_text and "SKU" in page_text:
        return "packing_slip"
    
    # Otherwise, use the indicator counts
    if packing_slip_count > label_count:
        return "packing_slip"
    else:
        return "label"

def group_label_with_packing_slips(pdf_document):
    """
    Group labels with their corresponding packing slips.
    Returns a list of dictionaries with label and packing slip information
    """
    groups = []
    i = 0
    total_pages = len(pdf_document)
    
    while i < total_pages:
        page = pdf_document[i]
        page_text = page.get_text()
        page_type = detect_page_type(page_text)
        
        if page_type == "label":
            # Found a label, now collect all following packing slips
            label_page = page
            label_index = i
            packing_slips = []
            packing_slip_indices = []
            
            # Look ahead for packing slips
            j = i + 1
            while j < total_pages:
                next_page = pdf_document[j]
                next_text = next_page.get_text()
                next_type = detect_page_type(next_text)
                
                if next_type == "packing_slip":
                    # Add this packing slip to the group
                    packing_slips.append(next_page)
                    packing_slip_indices.append(j)
                    j += 1
                else:
                    # Hit another label, stop collecting
                    break
            
            # Store the group
            if packing_slips:  # Only add if we found packing slips
                groups.append({
                    'label_page': label_page,
                    'label_index': label_index,
                    'packing_slips': packing_slips,
                    'packing_slip_indices': packing_slip_indices
                })
            
            # Move to the next unprocessed page
            i = j
        else:
            # Skip orphaned packing slips (shouldn't happen in normal cases)
            i += 1
    
    return groups

def add_page_number_with_slip_count(page, page_number, slip_number, total_slips):
    """
    Add page number with packing slip count for multi-slip orders.
    e.g., "Page: 42 (Slip 1 of 3)"
    """
    page_width = page.rect.width
    page_height = page.rect.height
    
    x = page_width - 120  # Adjusted for longer text
    y = page_height - 20
    
    text = f"Page: {page_number} (Slip {slip_number} of {total_slips})"
    page.insert_text((x, y), text,
                     fontname="Helvetica-Bold", fontsize=8, color=(0, 0, 0))

def extract_items(page_content):
    """
    Extract all items (product name, variation, SKU, quantity) from the given page content.
    Handles multi-line product names and variations.
    Optimized for better performance.
    """
    # Find the start of the product listing section
    product_section_start = page_content.find("Product Name")
    if product_section_start == -1:
        return []  # No product section found
    
    # Find the end of the product listing (Qty Total line)
    qty_total_position = page_content.find("Qty Total:", product_section_start)
    if qty_total_position == -1:
        qty_total_position = len(page_content)
    
    # Extract just the product section
    product_section = page_content[product_section_start:qty_total_position]
    
    # Find all SKUs and quantities using compiled regex for speed
    sku_qty_matches = list(SKU_QTY_PATTERN.finditer(product_section))
    
    items = []
    
    for i, match in enumerate(sku_qty_matches):
        sku = match.group(1)
        qty = int(match.group(2))
        
        # Find the start position for the product name
        if i == 0:
            # For the first item, start after "Product Name SKU Seller SKU Qty"
            header_end = product_section.find("Qty", product_section.find("Seller SKU"))
            if header_end == -1:
                continue
            name_start = header_end + 3  # Length of "Qty"
        else:
            # For subsequent items, start after the previous match
            name_start = sku_qty_matches[i-1].end()
        
        # Find the end position for the product name (just before "Default" or other variation)
        name_end = match.start()
        
        # Extract the raw product text
        product_text = product_section[name_start:name_end].strip()
        
        # Extract variation (text right before the SKU)
        variation = ""
        # Look for "Default" before the SKU
        default_pos = product_text.rfind("Default")
        if default_pos != -1:
            # "Default" is the variation
            variation = "Default"
            # Remove "Default" and get the actual product name
            product_name = product_text[:default_pos].strip()
        else:
            # If not "Default", check for any text in the last line as possible variation
            lines = product_text.split('\n')
            if lines:
                variation = lines[-1].strip()
                product_name = '\n'.join(lines[:-1]).strip() if len(lines) > 1 else product_text
                
                # If the last line contains the SKU pattern (might be captured in product_text), clean it
                if SKU_QTY_PATTERN.search(variation):
                    variation = ""
                    product_name = product_text
        
        # Clean up product name (remove any line with SKU/Seller SKU)
        product_name = '\n'.join(line for line in product_name.split('\n') 
                               if not any(kw in line for kw in ["SKU", "Seller"]))
        
        # Join multi-line product names with spaces
        product_name = ' '.join(product_name.split())
        
        items.append({
            "product_name": product_name,
            "variation": variation if variation != "Default" else "",  # Don't include "Default" as a variation
            "sku": sku,
            "qty": qty
        })
    
    return items

def extract_order_id(page_text):
    """
    Extract the order ID from page text.
    Optimized with compiled regex.
    """
    order_id_match = ORDER_ID_PATTERN.search(page_text)
    if order_id_match:
        return order_id_match.group(1)
    return None

def extract_tracking_number(page_text):
    """
    Extract the tracking number from page text.
    Optimized with compiled regex patterns.
    """
    for pattern in TRACKING_PATTERNS:
        tracking_match = pattern.search(page_text)
        if tracking_match:
            return tracking_match.group(1).strip()
    
    return None

def add_hazmat_image_to_page(page, image_path, image_width=232, image_height=100):
    """
    Add the hazmat image to the bottom center of a PDF page using PyMuPDF.
    """
    if not os.path.exists(image_path):
        return  # Skip if image doesn't exist
        
    x = (page.rect.width - image_width) / 7
    y = page.rect.height - image_height - 0.5  # 0.5 units from the bottom
    rect = fitz.Rect(x, y, x + image_width, y + image_height)
    page.insert_image(rect, filename=image_path)

def add_multi_qty_header(page, items):
    """
    Add a bold "MULTI-QUANTITY ORDER" alert at the bottom of packing slips with multiple quantities
    """
    if any(item['qty'] > 1 for item in items):
        # Position at the bottom of the page (1 inch from bottom)
        page_height = page.rect.height
        header_x = 50  # Left margin
        header_y = page_height - 72  # 1 inch (72 points) from bottom
        
        # Add header text in bold
        page.insert_text((header_x, header_y), "MULTI-QUANTITY ORDER",
                        fontname="Helvetica-Bold", fontsize=14, color=(0, 0, 0))
        
        # Underline the text
        text_width = 180  # Approximate width of the text
        page.draw_line((header_x, header_y + 5), (header_x + text_width, header_y + 5), 
                      width=1.5, color=(0, 0, 0))

def add_page_number(page, page_number):
    """
    Add page number to the bottom right corner of the page in bold
    Adjusted position and font size for better visibility
    """
    page_width = page.rect.width
    page_height = page.rect.height
    
    # Position at the bottom right of the page
    x = page_width - 65  
    y = page_height - 20  # 20 points from bottom edge
    
    # Add page number in bold with reduced font size (8 instead of 10)
    page.insert_text((x, y), f"Page: {page_number}",
                     fontname="Helvetica-Bold", fontsize=8, color=(0, 0, 0))

def add_label_count(page, sku, current_count, total_count):
    """
    Add label count (e.g., "SKU: T123 (1 of 5)") to packing slips for SKUs with qty >= 5
    Placed at the bottom left of the packing slip
    """
    page_width = page.rect.width
    page_height = page.rect.height
    
    # Position at the bottom left of the page
    x = 50  # 50 points from left edge
    y = page_height - 40  # 40 points from bottom edge
    
    # Add SKU count in bold
    page.insert_text((x, y), f"SKU: {sku} ({current_count} of {total_count})",
                     fontname="Helvetica-Bold", fontsize=10, color=(0, 0, 0))

def sort_orders(orders):
    """
    Sort orders according to the new sorting logic:
    
    - Warehouse (SKUs with 5+ occurrences, qty=1):
      * Hazmat first, then ground
      * Each sorted in descending order by frequency
    
    - Packing room (everything else):
      * Hazmat first, then non-hazmat
      * Within each category (hazmat/ground):
        1. Single item orders (1 item, qty=1)
        2. Single SKU orders (1 SKU, multiple qty)
        3. Multi-SKU orders (multiple items/SKUs)
      * Group identical orders together
      * Sort from most frequent to least
    
    Returns orders in two groups: warehouse and packingroom
    """
    # Count occurrences of each SKU across all orders with Qty Total = 1
    sku_occurrences = defaultdict(int)
    for order in orders:
        if order['qty_total'] == 1 and len(order['items']) == 1:
            sku_occurrences[order['items'][0]['sku']] += 1
    
    # Check for high-quantity SKUs (100+ occurrences)
    high_qty_skus = {}
    for sku, count in sku_occurrences.items():
        if count >= 100:
            # Find the first order with this SKU to get the product name
            for order in orders:
                if len(order['items']) == 1 and order['items'][0]['sku'] == sku:
                    high_qty_skus[sku] = {
                        'count': count,
                        'product_name': order['items'][0]['product_name'],
                        'is_hazmat': order.get('is_hazmat', False),
                        'orders': []
                    }
                    break
    
    # Create fingerprints for orders to identify identical ones
    order_fingerprints = {}
    for i, order in enumerate(orders):
        # Create a fingerprint based on SKUs and quantities
        items_info = sorted([(item['sku'], item['qty']) for item in order['items']])
        fingerprint = tuple(items_info)
        
        if fingerprint not in order_fingerprints:
            order_fingerprints[fingerprint] = []
        order_fingerprints[fingerprint].append(i)
    
    # Count frequency of each order fingerprint
    fingerprint_counts = {fp: len(indices) for fp, indices in order_fingerprints.items()}
    
    # Separate warehouse and packing room orders
    warehouse_hazmat = []
    warehouse_ground = []
    
    # For packing room, separate into categories
    packingroom_hazmat_single_item = []  # 1 item, qty=1
    packingroom_hazmat_single_sku = []   # 1 SKU, multiple qty
    packingroom_hazmat_multi_sku = []    # Multiple SKUs
    
    packingroom_ground_single_item = []  # 1 item, qty=1
    packingroom_ground_single_sku = []   # 1 SKU, multiple qty
    packingroom_ground_multi_sku = []    # Multiple SKUs
    
    # First, collect high quantity SKU orders if they exist
    for sku, info in high_qty_skus.items():
        high_qty_skus[sku]['orders'] = []
    
    # Sort orders into appropriate categories
    for i, order in enumerate(orders):
        # Skip duplicates to process them together with their first occurrence
        fingerprint = tuple(sorted([(item['sku'], item['qty']) for item in order['items']]))
        fingerprint_indices = order_fingerprints[fingerprint]
        if i != fingerprint_indices[0]:
            continue
            
        # Check if this is a high-quantity SKU order
        if (len(order['items']) == 1 and order['qty_total'] == 1 and 
            order['items'][0]['sku'] in high_qty_skus):
            sku = order['items'][0]['sku']
            # Collect all identical orders
            for idx in fingerprint_indices:
                high_qty_skus[sku]['orders'].append(orders[idx])
            continue
            
        # Check for warehouse eligibility (SKUs with 5+ occurrences, qty=1)
        if (len(order['items']) == 1 and order['qty_total'] == 1 and 
            sku_occurrences[order['items'][0]['sku']] >= 5):
            # Add all identical orders
            identical_orders = [orders[idx] for idx in fingerprint_indices]
            
            if order.get('is_hazmat', False):
                warehouse_hazmat.extend(identical_orders)
            else:
                warehouse_ground.extend(identical_orders)
        else:
            # Add to packing room
            identical_orders = [orders[idx] for idx in fingerprint_indices]
            
            if order.get('is_hazmat', False):
                # Categorize hazmat orders
                if len(order['items']) == 1:
                    if order['qty_total'] == 1:
                        packingroom_hazmat_single_item.extend(identical_orders)
                    else:
                        packingroom_hazmat_single_sku.extend(identical_orders)
                else:
                    packingroom_hazmat_multi_sku.extend(identical_orders)
            else:
                # Categorize ground orders
                if len(order['items']) == 1:
                    if order['qty_total'] == 1:
                        packingroom_ground_single_item.extend(identical_orders)
                    else:
                        packingroom_ground_single_sku.extend(identical_orders)
                else:
                    packingroom_ground_multi_sku.extend(identical_orders)
    
    # Sort warehouse orders by frequency (descending)
    warehouse_hazmat.sort(key=lambda x: -sku_occurrences[x['items'][0]['sku']])
    warehouse_ground.sort(key=lambda x: -sku_occurrences[x['items'][0]['sku']])
    
    # Function to get order frequency for sorting
    def get_order_frequency(order):
        fingerprint = tuple(sorted([(item['sku'], item['qty']) for item in order['items']]))
        return fingerprint_counts[fingerprint]
    
    # Sort packing room orders by frequency (descending) within each category
    packingroom_hazmat_single_item.sort(key=get_order_frequency, reverse=True)
    packingroom_hazmat_single_sku.sort(key=get_order_frequency, reverse=True)
    packingroom_hazmat_multi_sku.sort(key=get_order_frequency, reverse=True)
    
    packingroom_ground_single_item.sort(key=get_order_frequency, reverse=True)
    packingroom_ground_single_sku.sort(key=get_order_frequency, reverse=True)
    packingroom_ground_multi_sku.sort(key=get_order_frequency, reverse=True)
    
    # Combine all packing room orders in the correct order
    packingroom_hazmat = (packingroom_hazmat_single_item + 
                          packingroom_hazmat_single_sku + 
                          packingroom_hazmat_multi_sku)
    
    packingroom_ground = (packingroom_ground_single_item + 
                         packingroom_ground_single_sku + 
                         packingroom_ground_multi_sku)
    
    all_packingroom_orders = packingroom_hazmat + packingroom_ground
    
    warehouse_orders = warehouse_hazmat + warehouse_ground
    
    # Combine all orders for total count
    all_orders = warehouse_orders + all_packingroom_orders
    
    return {
        'warehouse': {
            'hazmat': warehouse_hazmat,
            'ground': warehouse_ground,
            'all': warehouse_orders
        },
        'packingroom': {
            'hazmat': {
                'single_item': packingroom_hazmat_single_item,
                'single_sku': packingroom_hazmat_single_sku,
                'multi_sku': packingroom_hazmat_multi_sku,
                'all': packingroom_hazmat
            },
            'ground': {
                'single_item': packingroom_ground_single_item,
                'single_sku': packingroom_ground_single_sku,
                'multi_sku': packingroom_ground_multi_sku,
                'all': packingroom_ground
            },
            'all': all_packingroom_orders
        },
        'high_qty_skus': high_qty_skus,
        'all': all_orders
    }

def create_warehouse_picklist_excel(warehouse_orders, output_path):
    """
    Create a properly formatted Excel file with the warehouse pick list
    
    Groups items by SKU and includes total quantities and page ranges.
    Now with improved formatting:
    - Includes variation in product name
    - Removed check box column
    - Qty comes after item name
    - Alternating row colors
    - Centered and bold qty values
    - Added "Picked by" and "Packed by" columns for employee tracking
    - Optimized for letter-size printing
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Warehouse Pick List"
    
    # Set column widths for letter size paper (8.5" x 11") with new columns
    worksheet.column_dimensions['A'].width = 45  # Item Name (slightly reduced to fit new columns)
    worksheet.column_dimensions['B'].width = 8   # Qty (centered)
    worksheet.column_dimensions['C'].width = 12  # SKU
    worksheet.column_dimensions['D'].width = 12  # Page Numbers
    worksheet.column_dimensions['E'].width = 15  # Picked by
    worksheet.column_dimensions['F'].width = 15  # Packed by
    
    # Create fonts
    title_font = Font(name="Arial", size=16, bold=True)
    header_font = Font(name="Arial", size=14, bold=True)
    content_font = Font(name="Arial", size=13)
    product_name_font = Font(name="Arial", size=13)
    qty_font = Font(name="Arial", size=13, bold=True)  # Bold for qty
    
    # Create fills for alternating rows
    light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Add HAZMAT PICK LIST header (merged cells A1-F1)
    worksheet.merge_cells('A1:F1')
    worksheet['A1'] = "HAZMAT PICK LIST"
    worksheet['A1'].font = title_font
    worksheet['A1'].alignment = Alignment(horizontal='center')
    
    # Add today's date (merged cells A2-F2)
    worksheet.merge_cells('A2:F2')
    worksheet['A2'] = datetime.datetime.now().strftime("%Y-%m-%d")
    worksheet['A2'].font = content_font
    worksheet['A2'].alignment = Alignment(horizontal='center')
    
    # Add column headers
    headers = ["Item Name", "Qty", "SKU", "Pages", "Picked by", "Packed by"]
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=3, column=col, value=header)
        cell.font = header_font
        # Center specific headers
        if header in ["Qty", "Picked by", "Packed by"]:
            cell.alignment = Alignment(horizontal='center')
    
    # Add border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Process hazmat items first - Group by SKU
    row_idx = 4  # Start after headers
    
    # Group hazmat items by SKU with page number tracking
    hazmat_items_by_sku = {}
    
    if warehouse_orders['hazmat']:
        for order in warehouse_orders['hazmat']:
            for item in order['items']:
                sku = item['sku']
                if sku not in hazmat_items_by_sku:
                    # Include variation in product name if present
                    product_name = item['product_name']
                    variation = item.get('variation', '')
                    if variation:
                        product_name = f"{variation} {product_name}"
                        
                    hazmat_items_by_sku[sku] = {
                        'product_name': product_name,
                        'quantity': 0,
                        'pages': []
                    }
                # Add to total quantity
                hazmat_items_by_sku[sku]['quantity'] += item['qty']
                # Add page number to list
                if 'page_number' in order:
                    hazmat_items_by_sku[sku]['pages'].append(order['page_number'])
        
        # Add grouped hazmat items to the worksheet
        row_counter = 0  # For alternating row colors
        for sku, data in hazmat_items_by_sku.items():
            # Apply alternating row background
            if row_counter % 2 == 1:  # Odd rows get shaded
                for col in range(1, 7):  # Updated to include new columns
                    worksheet.cell(row=row_idx, column=col).fill = light_fill
            
            # Sort page numbers and create ranges
            pages = sorted(data['pages'])
            page_ranges = []
            
            if pages:
                start = pages[0]
                prev = start
                
                for i in range(1, len(pages)):
                    if pages[i] != prev + 1:
                        if start == prev:
                            page_ranges.append(str(start))
                        else:
                            page_ranges.append(f"{start}-{prev}")
                        start = pages[i]
                    prev = pages[i]
                
                # Add the last range
                if start == prev:
                    page_ranges.append(str(start))
                else:
                    page_ranges.append(f"{start}-{prev}")
                
                page_text = ", ".join(page_ranges)
            else:
                page_text = ""
            
            # Add row to worksheet - Name, Qty, SKU, Pages, Picked by, Packed by
            name_cell = worksheet.cell(row=row_idx, column=1, value=data['product_name'])
            name_cell.font = product_name_font
            name_cell.border = thin_border
            name_cell.alignment = Alignment(wrap_text=True, vertical='center')
            
            qty_cell = worksheet.cell(row=row_idx, column=2, value=data['quantity'])
            qty_cell.font = qty_font
            qty_cell.border = thin_border
            qty_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            sku_cell = worksheet.cell(row=row_idx, column=3, value=sku)
            sku_cell.font = content_font
            sku_cell.border = thin_border
            sku_cell.alignment = Alignment(vertical='center')
            
            page_cell = worksheet.cell(row=row_idx, column=4, value=page_text)
            page_cell.font = content_font
            page_cell.border = thin_border
            page_cell.alignment = Alignment(vertical='center')
            
            # New columns for employee tracking
            picked_cell = worksheet.cell(row=row_idx, column=5, value="")
            picked_cell.font = content_font
            picked_cell.border = thin_border
            picked_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            packed_cell = worksheet.cell(row=row_idx, column=6, value="")
            packed_cell.font = content_font
            packed_cell.border = thin_border
            packed_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            row_idx += 1
            row_counter += 1
    
    # Add GROUND PICK LIST header
    row_idx += 1  # Add a blank row
    worksheet.merge_cells(f'A{row_idx}:F{row_idx}')
    worksheet[f'A{row_idx}'] = "GROUND PICK LIST"
    worksheet[f'A{row_idx}'].font = title_font
    worksheet[f'A{row_idx}'].alignment = Alignment(horizontal='center')
    row_idx += 1
    
    # Group ground items by SKU with page number tracking
    ground_items_by_sku = {}
    
    if warehouse_orders['ground']:
        for order in warehouse_orders['ground']:
            for item in order['items']:
                sku = item['sku']
                if sku not in ground_items_by_sku:
                    # Include variation in product name if present
                    product_name = item['product_name']
                    variation = item.get('variation', '')
                    if variation:
                        product_name = f"{variation} {product_name}"
                        
                    ground_items_by_sku[sku] = {
                        'product_name': product_name,
                        'quantity': 0,
                        'pages': []
                    }
                # Add to total quantity
                ground_items_by_sku[sku]['quantity'] += item['qty']
                # Add page number to list
                if 'page_number' in order:
                    ground_items_by_sku[sku]['pages'].append(order['page_number'])
        
        # Add grouped ground items to the worksheet
        row_counter = 0  # Reset counter for ground section
        for sku, data in ground_items_by_sku.items():
            # Apply alternating row background
            if row_counter % 2 == 1:  # Odd rows get shaded
                for col in range(1, 7):  # Updated to include new columns
                    worksheet.cell(row=row_idx, column=col).fill = light_fill
            
            # Sort page numbers and create ranges
            pages = sorted(data['pages'])
            page_ranges = []
            
            if pages:
                start = pages[0]
                prev = start
                
                for i in range(1, len(pages)):
                    if pages[i] != prev + 1:
                        if start == prev:
                            page_ranges.append(str(start))
                        else:
                            page_ranges.append(f"{start}-{prev}")
                        start = pages[i]
                    prev = pages[i]
                
                # Add the last range
                if start == prev:
                    page_ranges.append(str(start))
                else:
                    page_ranges.append(f"{start}-{prev}")
                
                page_text = ", ".join(page_ranges)
            else:
                page_text = ""
            
            # Add row to worksheet - Name, Qty, SKU, Pages, Picked by, Packed by
            name_cell = worksheet.cell(row=row_idx, column=1, value=data['product_name'])
            name_cell.font = product_name_font
            name_cell.border = thin_border
            name_cell.alignment = Alignment(wrap_text=True, vertical='center')
            
            qty_cell = worksheet.cell(row=row_idx, column=2, value=data['quantity'])
            qty_cell.font = qty_font
            qty_cell.border = thin_border
            qty_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            sku_cell = worksheet.cell(row=row_idx, column=3, value=sku)
            sku_cell.font = content_font
            sku_cell.border = thin_border
            sku_cell.alignment = Alignment(vertical='center')
            
            page_cell = worksheet.cell(row=row_idx, column=4, value=page_text)
            page_cell.font = content_font
            page_cell.border = thin_border
            page_cell.alignment = Alignment(vertical='center')
            
            # New columns for employee tracking
            picked_cell = worksheet.cell(row=row_idx, column=5, value="")
            picked_cell.font = content_font
            picked_cell.border = thin_border
            picked_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            packed_cell = worksheet.cell(row=row_idx, column=6, value="")
            packed_cell.font = content_font
            packed_cell.border = thin_border
            packed_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            row_idx += 1
            row_counter += 1
    
    # Set row heights to accommodate wrapped text
    for row in range(4, row_idx):
        worksheet.row_dimensions[row].height = 35  # Slightly increased for new columns
    
    # Save the workbook
    workbook.save(output_path)

def save_sku_counts_to_csv(packingroom_orders, csv_output_path):
    """
    Save SKU counts to a CSV file with sections matching the new packingroom order categories:
    
    1. HAZMAT - SINGLE ITEM ORDERS (SKU, Product Name, Variation, Quantity)
    2. HAZMAT - SINGLE SKU ORDERS (MULTI QTY)
    3. HAZMAT - MULTI SKU ORDERS
    4. GROUND - SINGLE ITEM ORDERS
    5. GROUND - SINGLE SKU ORDERS (MULTI QTY)
    6. GROUND - MULTI SKU ORDERS
    """
    # Create dictionaries for each section
    hazmat_single_item_counts = defaultdict(lambda: {"product_name": "", "variation": "", "quantity": 0})
    hazmat_single_sku_counts = defaultdict(lambda: {"product_name": "", "variation": "", "quantity": 0})
    hazmat_multi_sku_counts = defaultdict(lambda: {"product_name": "", "variation": "", "quantity": 0})
    
    ground_single_item_counts = defaultdict(lambda: {"product_name": "", "variation": "", "quantity": 0})
    ground_single_sku_counts = defaultdict(lambda: {"product_name": "", "variation": "", "quantity": 0})
    ground_multi_sku_counts = defaultdict(lambda: {"product_name": "", "variation": "", "quantity": 0})

    # First pass: collect all product names and variations for each SKU (we'll choose the most complete ones)
    all_product_names = defaultdict(list)
    all_variations = defaultdict(list)
    
    # Use the dictionary structure from the new sorting
    for order in packingroom_orders['all']:
        for item in order['items']:
            all_product_names[item['sku']].append(item['product_name'])
            if 'variation' in item:
                all_variations[item['sku']].append(item['variation'])
    
    # Select the most informative product name for each SKU (usually the longest)
    best_product_names = {}
    best_variations = {}
    
    for sku, names in all_product_names.items():
        if names:
            # Filter out empty names
            valid_names = [name for name in names if name.strip()]
            if valid_names:
                # Choose the longest name as it's likely the most complete
                best_product_names[sku] = max(valid_names, key=len)
            else:
                best_product_names[sku] = ""
    
    for sku, variations in all_variations.items():
        if variations:
            # Filter out empty variations
            valid_variations = [var for var in variations if var.strip()]
            if valid_variations:
                # Choose the first non-empty variation (they should be consistent for a SKU)
                best_variations[sku] = valid_variations[0]
            else:
                best_variations[sku] = ""
        else:
            best_variations[sku] = ""
    
    # Second pass: categorize and count total quantities for each SKU by category
    # Hazmat - Single Item
    for order in packingroom_orders['hazmat']['single_item']:
        for item in order['items']:
            hazmat_single_item_counts[item['sku']]["product_name"] = best_product_names.get(item['sku'], "")
            hazmat_single_item_counts[item['sku']]["variation"] = best_variations.get(item['sku'], "")
            hazmat_single_item_counts[item['sku']]["quantity"] += item['qty']
    
    # Hazmat - Single SKU, Multiple Qty
    for order in packingroom_orders['hazmat']['single_sku']:
        for item in order['items']:
            hazmat_single_sku_counts[item['sku']]["product_name"] = best_product_names.get(item['sku'], "")
            hazmat_single_sku_counts[item['sku']]["variation"] = best_variations.get(item['sku'], "")
            hazmat_single_sku_counts[item['sku']]["quantity"] += item['qty']
    
    # Hazmat - Multi SKU
    for order in packingroom_orders['hazmat']['multi_sku']:
        for item in order['items']:
            hazmat_multi_sku_counts[item['sku']]["product_name"] = best_product_names.get(item['sku'], "")
            hazmat_multi_sku_counts[item['sku']]["variation"] = best_variations.get(item['sku'], "")
            hazmat_multi_sku_counts[item['sku']]["quantity"] += item['qty']
    
    # Ground - Single Item
    for order in packingroom_orders['ground']['single_item']:
        for item in order['items']:
            ground_single_item_counts[item['sku']]["product_name"] = best_product_names.get(item['sku'], "")
            ground_single_item_counts[item['sku']]["variation"] = best_variations.get(item['sku'], "")
            ground_single_item_counts[item['sku']]["quantity"] += item['qty']
    
    # Ground - Single SKU, Multiple Qty
    for order in packingroom_orders['ground']['single_sku']:
        for item in order['items']:
            ground_single_sku_counts[item['sku']]["product_name"] = best_product_names.get(item['sku'], "")
            ground_single_sku_counts[item['sku']]["variation"] = best_variations.get(item['sku'], "")
            ground_single_sku_counts[item['sku']]["quantity"] += item['qty']
    
    # Ground - Multi SKU
    for order in packingroom_orders['ground']['multi_sku']:
        for item in order['items']:
            ground_multi_sku_counts[item['sku']]["product_name"] = best_product_names.get(item['sku'], "")
            ground_multi_sku_counts[item['sku']]["variation"] = best_variations.get(item['sku'], "")
            ground_multi_sku_counts[item['sku']]["quantity"] += item['qty']
            
    # Create a single CSV file with all sections
    with open(csv_output_path, mode='w', newline='') as file:
        writer = csv.writer(file)
        
        # HAZMAT sections
        writer.writerow(["HAZMAT PICK LIST"])
        
        # HAZMAT - Single Item Orders
        writer.writerow(["SINGLE ITEM ORDERS"])
        writer.writerow(["SKU", "Product Name", "Variation", "Quantity"])
        
        sorted_hazmat_single_item = sorted(hazmat_single_item_counts.items(), key=lambda x: (-x[1]["quantity"], x[0]))
        for sku, data in sorted_hazmat_single_item:
            writer.writerow([sku, data["product_name"], data["variation"], data["quantity"]])
        
        writer.writerow([])
        
        # HAZMAT - Single SKU Orders (Multi Qty)
        writer.writerow(["SINGLE SKU ORDERS (MULTI QTY)"])
        writer.writerow(["SKU", "Product Name", "Variation", "Quantity"])
        
        sorted_hazmat_single_sku = sorted(hazmat_single_sku_counts.items(), key=lambda x: (-x[1]["quantity"], x[0]))
        for sku, data in sorted_hazmat_single_sku:
            writer.writerow([sku, data["product_name"], data["variation"], data["quantity"]])
        
        writer.writerow([])
        
        # HAZMAT - Multi SKU Orders
        writer.writerow(["MULTI SKU ORDERS"])
        writer.writerow(["SKU", "Product Name", "Variation", "Quantity"])
        
        sorted_hazmat_multi_sku = sorted(hazmat_multi_sku_counts.items(), key=lambda x: (-x[1]["quantity"], x[0]))
        for sku, data in sorted_hazmat_multi_sku:
            writer.writerow([sku, data["product_name"], data["variation"], data["quantity"]])
        
        writer.writerow([])
        
        # GROUND sections
        writer.writerow(["GROUND PICK LIST"])
        
        # GROUND - Single Item Orders
        writer.writerow(["SINGLE ITEM ORDERS"])
        writer.writerow(["SKU", "Product Name", "Variation", "Quantity"])
        
        sorted_ground_single_item = sorted(ground_single_item_counts.items(), key=lambda x: (-x[1]["quantity"], x[0]))
        for sku, data in sorted_ground_single_item:
            writer.writerow([sku, data["product_name"], data["variation"], data["quantity"]])
        
        writer.writerow([])
        
        # GROUND - Single SKU Orders (Multi Qty)
        writer.writerow(["SINGLE SKU ORDERS (MULTI QTY)"])
        writer.writerow(["SKU", "Product Name", "Variation", "Quantity"])
        
        sorted_ground_single_sku = sorted(ground_single_sku_counts.items(), key=lambda x: (-x[1]["quantity"], x[0]))
        for sku, data in sorted_ground_single_sku:
            writer.writerow([sku, data["product_name"], data["variation"], data["quantity"]])
        
        writer.writerow([])
        
        # GROUND - Multi SKU Orders
        writer.writerow(["MULTI SKU ORDERS"])
        writer.writerow(["SKU", "Product Name", "Variation", "Quantity"])
        
        sorted_ground_multi_sku = sorted(ground_multi_sku_counts.items(), key=lambda x: (-x[1]["quantity"], x[0]))
        for sku, data in sorted_ground_multi_sku:
            writer.writerow([sku, data["product_name"], data["variation"], data["quantity"]])

def move_to_discard_folder(original_file):
    """
    Move the original file to a DISCARD folder after processing.
    """
    discard_dir = os.path.join(os.path.dirname(original_file), "DISCARD")
    os.makedirs(discard_dir, exist_ok=True)
    discard_path = os.path.join(discard_dir, os.path.basename(original_file))
    os.rename(original_file, discard_path)

def process_pdfs(input_folder, output_folder, hazmat_keywords, auto_open=False, status_callback=None, hazmat_sticker_enabled=True):
    """
    Process PDFs with support for multiple packing slips per label.
    
    Args:
        input_folder: Path to folder containing PDF files
        output_folder: Path to save processed files
        hazmat_keywords: List of keywords to identify hazmat items
        auto_open: If True, automatically open output PDFs (disabled in web version)
        status_callback: Function to call with status updates
        hazmat_sticker_enabled: If True, add hazmat stickers to qualifying labels
    
    Returns:
        dict: Processing results and statistics
    """
    start_time = time.time()  # Start timing the process
    
    if status_callback:
        status_callback(f"Looking for PDF files in {input_folder}")
    
    pdf_files = [f for f in os.listdir(input_folder) if f.lower().endswith(".pdf")]
    if not pdf_files:
        if status_callback:
            status_callback("No PDF files found in the input folder.")
        return {'success': False}
    
    # Create timestamp for output directory
    timestamp = datetime.datetime.now().strftime("%I-%M_%m-%d-%y")
    output_dir = os.path.join(output_folder, "Output", timestamp)
    os.makedirs(output_dir, exist_ok=True)
    
    # Get hazmat image path
    import sys
    if getattr(sys, 'frozen', False):
        # Running as compiled application
        application_path = sys._MEIPASS
    else:
        # Running as script
        application_path = os.path.dirname(os.path.abspath(__file__))
    
    hazmat_image_path = os.path.join(application_path, "hazmat.png")
    
    if not os.path.exists(hazmat_image_path):
        if status_callback:
            status_callback(f"Warning: Hazmat image not found at {hazmat_image_path}")
    
    # Collect all orders from all PDFs
    all_orders = []
    processed_files = []
    all_tracking_numbers = set()
    duplicate_tracking_numbers = set()
    duplicate_details = []
    total_orders = 0
    multi_slip_orders = 0
    max_slips_per_order = 1
    
    if status_callback:
        status_callback(f"Found {len(pdf_files)} PDF files to process.")
    
    for pdf_file in pdf_files:
        input_pdf_path = os.path.join(input_folder, pdf_file)
        if status_callback:
            status_callback(f"Processing: {pdf_file}")
        
        # Open the PDF
        pdf_document = fitz.open(input_pdf_path)
        total_pages = len(pdf_document)
        if status_callback:
            status_callback(f"  - {total_pages} pages found in {pdf_file}")
        
        # Try to detect if this PDF has multi-slip orders
        label_groups = group_label_with_packing_slips(pdf_document)
        
        if label_groups:
            # Process using multi-slip logic
            if status_callback:
                status_callback(f"  - Using multi-slip processing for {pdf_file}")
            
            for group in label_groups:
                label_page = group['label_page']
                packing_slips = group['packing_slips']
                
                # Track multi-slip statistics
                if len(packing_slips) > 1:
                    multi_slip_orders += 1
                    max_slips_per_order = max(max_slips_per_order, len(packing_slips))
                
                # Extract information from all packing slips
                all_items = []
                qty_total = 0
                order_id = None
                tracking_number = None
                is_hazmat = False
                
                # Process each packing slip in the group
                for i, packing_slip_page in enumerate(packing_slips):
                    page_text = packing_slip_page.get_text()
                    
                    # Extract order ID and tracking number from first packing slip
                    if order_id is None:
                        order_id = extract_order_id(page_text)
                    if tracking_number is None:
                        tracking_number = extract_tracking_number(label_page.get_text())
                    
                    # Extract items from this packing slip
                    items = extract_items(page_text)
                    all_items.extend(items)
                    
                    # Extract qty total using compiled regex
                    qty_total_match = QTY_TOTAL_PATTERN.search(page_text)
                    if qty_total_match:
                        qty_total += int(qty_total_match.group(1))
                    
                    # Check for hazmat keywords
                    if any(keyword.lower() in page_text.lower() for keyword in hazmat_keywords):
                        is_hazmat = True
                    
                    # Add multi-quantity header if needed
                    add_multi_qty_header(packing_slip_page, items)
                
                # Check for duplicate tracking numbers
                if tracking_number:
                    if tracking_number in all_tracking_numbers:
                        duplicate_tracking_numbers.add(tracking_number)
                        duplicate_details.append(f"Tracking: {tracking_number}, Order: {order_id}")
                        if status_callback:
                            status_callback(f"  - Found duplicate tracking number: {tracking_number}")
                    all_tracking_numbers.add(tracking_number)
                    total_orders += 1
                
                # Add hazmat image to label if needed and if stickers are enabled
                if is_hazmat and hazmat_sticker_enabled:
                    if status_callback:
                        status_callback(f"  - Hazmat keyword found, adding sticker to label")
                    add_hazmat_image_to_page(label_page, hazmat_image_path)
                elif is_hazmat and not hazmat_sticker_enabled:
                    if status_callback:
                        status_callback(f"  - Hazmat keyword found, but stickers disabled")
                
                # Add to all orders
                order_info = {
                    'items': all_items,
                    'qty_total': qty_total,
                    'pages': (label_page, packing_slips),
                    'source_pdf': pdf_document,
                    'page_numbers': (group['label_index'], group['packing_slip_indices']),
                    'is_hazmat': is_hazmat,
                    'order_id': order_id,
                    'tracking_number': tracking_number,
                    'num_packing_slips': len(packing_slips)
                }
                
                all_orders.append(order_info)
        else:
            # Fall back to traditional processing (every 2 pages)
            if status_callback:
                status_callback(f"  - Using traditional 2-page processing for {pdf_file}")
                
            for i in range(0, total_pages, 2):
                label_page = pdf_document[i]
                if i + 1 < total_pages:
                    packing_slip_page = pdf_document[i + 1]
                    page_text = packing_slip_page.get_text()
                    
                    # Extract order ID for reference
                    order_id = extract_order_id(page_text)
                    
                    # Extract tracking number and check for duplicates
                    tracking_number = extract_tracking_number(label_page.get_text())
                    
                    if tracking_number:
                        # Check if it's a duplicate
                        if tracking_number in all_tracking_numbers:
                            duplicate_tracking_numbers.add(tracking_number)
                            duplicate_details.append(f"Tracking: {tracking_number}, Order: {order_id}")
                            if status_callback:
                                status_callback(f"Found duplicate tracking number: {tracking_number} (Order ID: {order_id})")
                        all_tracking_numbers.add(tracking_number)
                        total_orders += 1  # Count each tracking number as an order
                    
                    items = extract_items(page_text)
                    qty_total_match = QTY_TOTAL_PATTERN.search(page_text)
                    qty_total = int(qty_total_match.group(1)) if qty_total_match else 0
                    
                    # Add multi-quantity header if needed
                    add_multi_qty_header(packing_slip_page, items)
                    
                    # Check for hazmat keywords
                    is_hazmat = any(keyword.lower() in page_text.lower() for keyword in hazmat_keywords)
                    
                    if is_hazmat and hazmat_sticker_enabled:
                        if status_callback:
                            status_callback(f"  - Hazmat keyword found on page {i + 2}, adding sticker to label.")
                        add_hazmat_image_to_page(label_page, hazmat_image_path)
                    elif is_hazmat and not hazmat_sticker_enabled:
                        if status_callback:
                            status_callback(f"  - Hazmat keyword found on page {i + 2}, but stickers disabled.")
                    
                    # Add to all orders with reference to original PDF and pages
                    order_info = {
                        'items': items,
                        'qty_total': qty_total,
                        'pages': (label_page, [packing_slip_page]),  # Wrap in list for consistency
                        'source_pdf': pdf_document,
                        'page_numbers': (i, [i + 1]),  # Wrap in list for consistency
                        'is_hazmat': is_hazmat,
                        'order_id': order_id,
                        'tracking_number': tracking_number,
                        'num_packing_slips': 1
                    }
                    
                    all_orders.append(order_info)
        
        processed_files.append(input_pdf_path)
    
    # Sort all orders according to the new rules
    if status_callback:
        status_callback(f"Sorting {len(all_orders)} orders...")
    
    sorted_orders = sort_orders(all_orders)
    
    # Output file paths
    output_files = []
    
    # Process high-quantity SKUs - REMOVED tkinter messagebox for web compatibility
    high_qty_skus = sorted_orders['high_qty_skus']
    for sku, info in high_qty_skus.items():
        if info['count'] >= 100:
            if status_callback:
                status_callback(f"Found high-quantity SKU: {sku} - {info['product_name']} ({info['count']} orders)")
            
            # In web version, automatically create separate file (no user prompt)
            if status_callback:
                status_callback(f"Creating separate file for high-quantity SKU: {sku}")
            
            # Create separate PDF for this SKU
            sku_pdf_path = os.path.join(output_dir, f"sku_{sku}_{info['count']}.pdf")
            output_files.append(sku_pdf_path)
            
            sku_pdf = fitz.open()
            
            # Add page numbers and label counts to orders
            page_number = 1
            order_count = 1
            total_count = len(info['orders'])
            
            # Add all orders for this SKU to the PDF
            for order in info['orders']:
                # Add page number to the order for referencing in the picklist
                order['page_number'] = page_number
                
                # Add pages to the PDF
                sku_pdf.insert_pdf(order['source_pdf'], 
                                  from_page=order['page_numbers'][0], 
                                  to_page=order['page_numbers'][0])
                
                # Add all packing slips
                for j, slip_index in enumerate(order['page_numbers'][1]):
                    slip_page = order['source_pdf'][slip_index]
                    # Add page number with slip count if multiple slips
                    if len(order['page_numbers'][1]) > 1:
                        add_page_number_with_slip_count(slip_page, page_number, j + 1, len(order['page_numbers'][1]))
                    else:
                        add_page_number(slip_page, page_number)
                    
                    # Add SKU count on the first packing slip
                    if j == 0:
                        add_label_count(slip_page, sku, order_count, total_count)
                    
                    sku_pdf.insert_pdf(order['source_pdf'], 
                                      from_page=slip_index, 
                                      to_page=slip_index)
                
                # Increment counters
                page_number += 1
                order_count += 1
            
            sku_pdf.save(sku_pdf_path)
            sku_pdf.close()
            
            if status_callback:
                status_callback(f"Created separate file for SKU {sku}: {sku_pdf_path}")
            
            # Remove these orders from warehouse processing
            if info['is_hazmat']:
                sorted_orders['warehouse']['hazmat'] = [
                    order for order in sorted_orders['warehouse']['hazmat']
                    if len(order['items']) != 1 or order['items'][0]['sku'] != sku
                ]
            else:
                sorted_orders['warehouse']['ground'] = [
                    order for order in sorted_orders['warehouse']['ground']
                    if len(order['items']) != 1 or order['items'][0]['sku'] != sku
                ]
            
            # Update the all list
            sorted_orders['warehouse']['all'] = sorted_orders['warehouse']['hazmat'] + sorted_orders['warehouse']['ground']
    
    # Process warehouse orders
    warehouse_pdf_path = os.path.join(output_dir, "warehouse_labels.pdf")
    output_files.append(warehouse_pdf_path)
    
    # Count SKUs for label counting (e.g., "1 of 5", "2 of 5")
    sku_counts = defaultdict(int)
    for order in sorted_orders['warehouse']['all']:
        if len(order['items']) == 1:
            sku_counts[order['items'][0]['sku']] += 1
    
    # Add page numbers and label counts to warehouse orders
    page_number = 1
    sku_counters = defaultdict(int)
    
    # Create warehouse PDF with both hazmat and ground orders
    if sorted_orders['warehouse']['all']:
        warehouse_pdf = fitz.open()
        
        # Process all warehouse orders
        for order in sorted_orders['warehouse']['all']:
            # Add page number to the order for referencing in the picklist
            order['page_number'] = page_number
            
            # Add the order to PDF
            warehouse_pdf.insert_pdf(order['source_pdf'], 
                                   from_page=order['page_numbers'][0], 
                                   to_page=order['page_numbers'][0])
            
            # Add all packing slips with page numbers
            for j, slip_index in enumerate(order['page_numbers'][1]):
                slip_page = order['source_pdf'][slip_index]
                
                # Add page number with slip count if multiple slips
                if len(order['page_numbers'][1]) > 1:
                    add_page_number_with_slip_count(slip_page, page_number, j + 1, len(order['page_numbers'][1]))
                else:
                    add_page_number(slip_page, page_number)
                
                # Add SKU count for SKUs with qty >= 5 (on first slip only)
                if j == 0 and len(order['items']) == 1:
                    sku = order['items'][0]['sku']
                    if sku_counts[sku] >= 5:
                        sku_counters[sku] += 1
                        current_count = sku_counters[sku]
                        total_count = sku_counts[sku]
                        add_label_count(slip_page, sku, current_count, total_count)
                
                warehouse_pdf.insert_pdf(order['source_pdf'], 
                                       from_page=slip_index, 
                                       to_page=slip_index)
            
            # Increment page number for the next order
            page_number += 1
        
        warehouse_pdf.save(warehouse_pdf_path)
        warehouse_pdf.close()
        
        if status_callback:
            status_callback(f"Warehouse labels saved to {warehouse_pdf_path}")
        
        # Create warehouse pick list Excel file
        picklist_path = os.path.join(output_dir, "warehouse_picklist.xlsx")
        create_warehouse_picklist_excel(sorted_orders['warehouse'], picklist_path)
        
        if status_callback:
            status_callback(f"Warehouse pick list saved to {picklist_path}")
    
    # Process packingroom orders
    packingroom_pdf_path = os.path.join(output_dir, "packingroom_labels.pdf")
    output_files.append(packingroom_pdf_path)
    
    # Add page numbers to packingroom orders
    page_number = 1
    
    if sorted_orders['packingroom']['all']:
        packingroom_pdf = fitz.open()
        
        # Process all packingroom orders
        for order in sorted_orders['packingroom']['all']:
            # Add page number to the order for referencing
            order['page_number'] = page_number
            
            # Add label
            packingroom_pdf.insert_pdf(order['source_pdf'], 
                                     from_page=order['page_numbers'][0], 
                                     to_page=order['page_numbers'][0])
            
            # Add all packing slips with page numbers
            for j, slip_index in enumerate(order['page_numbers'][1]):
                slip_page = order['source_pdf'][slip_index]
                
                # Add page number with slip count if multiple slips
                if len(order['page_numbers'][1]) > 1:
                    add_page_number_with_slip_count(slip_page, page_number, j + 1, len(order['page_numbers'][1]))
                else:
                    add_page_number(slip_page, page_number)
                
                packingroom_pdf.insert_pdf(order['source_pdf'], 
                                         from_page=slip_index, 
                                         to_page=slip_index)
            
            # Increment page number for the next order
            page_number += 1
        
        packingroom_pdf.save(packingroom_pdf_path)
        packingroom_pdf.close()
        
        if status_callback:
            status_callback(f"Packingroom labels saved to {packingroom_pdf_path}")
        
        # Create pick list CSV for packingroom orders
        csv_output_path = os.path.join(output_dir, "packingroom_pick_list.csv")
        save_sku_counts_to_csv(sorted_orders['packingroom'], csv_output_path)
        
        if status_callback:
            status_callback(f"Packingroom pick list saved to {csv_output_path}")
    
    # Close all source PDFs
    for order in all_orders:
        if 'source_pdf' in order:
            try:
                if not order['source_pdf'].is_closed:
                    order['source_pdf'].close()
            except:
                pass  # Already closed
    
    # Skip moving files to DISCARD folder in web version (files are temporary)
    
    # Skip auto-open in web version (not supported)
    
    # Calculate processing time
    end_time = time.time()
    processing_time = end_time - start_time
    
    # Format processing time based on duration
    if processing_time < 60:
        time_str = f"{processing_time:.1f} seconds"
    else:
        minutes = int(processing_time // 60)
        seconds = processing_time % 60
        time_str = f"{minutes} min {seconds:.1f} sec"
    
    # Print summary statistics
    if status_callback:
        status_callback("\n--- ORDER SUMMARY ---")
        status_callback(f"Total Orders: {total_orders}")
        if multi_slip_orders > 0:
            status_callback(f"Multi-slip Orders: {multi_slip_orders}")
            status_callback(f"Max Slips per Order: {max_slips_per_order}")
        status_callback(f"Duplicate Orders: {len(duplicate_tracking_numbers)}")
        status_callback(f"Processing Time: {time_str}")
        status_callback(f"Warehouse Orders: {len(sorted_orders['warehouse']['all'])}")
        status_callback(f"  - Hazmat: {len(sorted_orders['warehouse']['hazmat'])}")
        status_callback(f"  - Ground: {len(sorted_orders['warehouse']['ground'])}")
        status_callback(f"Packingroom Orders: {len(sorted_orders['packingroom']['all'])}")
        status_callback(f"  - Hazmat: {len(sorted_orders['packingroom']['hazmat']['all'])}")
        status_callback(f"  - Ground: {len(sorted_orders['packingroom']['ground']['all'])}")
    
    # Return processing results and counters
    return {
        'success': True,
        'total_orders': total_orders,
        'duplicate_orders': len(duplicate_tracking_numbers),
        'duplicate_details': duplicate_details,
        'processing_time': time_str,
        'multi_slip_orders': multi_slip_orders,
        'max_slips_per_order': max_slips_per_order
    }